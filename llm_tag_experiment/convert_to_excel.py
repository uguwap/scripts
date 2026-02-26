"""
Конвертирует CSV-результат LLM-классификации в человекочитаемый Excel.

Лист 1 — «Транзакции»:  все транзакции с тегами LLM + существующим SQL-тегом.
Лист 2 — «Сводка по категориям»:  количество и сумма по level1 → level2 → level3.
Лист 3 — «Сравнение SQL vs LLM»:  как текущий SQL-тег соотносится с LLM-классификацией.

Запуск:
  python convert_to_excel.py                         # берёт result_adaptive.csv
  python convert_to_excel.py path/to/other.csv       # произвольный CSV
"""
import csv
import os
import sys
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Устанавливаю openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Пути ──────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

if len(sys.argv) > 1:
    CSV_PATH = sys.argv[1]
else:
    CSV_PATH = os.path.join(SCRIPT_DIR, "result_adaptive.csv")

XLSX_PATH = os.path.splitext(CSV_PATH)[0] + ".xlsx"

# ── Стили ─────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
INCOMING_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
OUTGOING_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# ── Русские заголовки ─────────────────────────────────────
RU_HEADERS = {
    "document_number": "No документа",
    "document_operation_date": "Дата операции",
    "payer_or_recipient_name": "Контрагент",
    "payer_or_recipient_inn": "ИНН контрагента",
    "debit_amount": "Дебет (списание)",
    "credit_amount": "Кредит (поступление)",
    "payment_purpose": "Назначение платежа",
    "existing_tag": "Текущий тег (SQL)",
    "is_leasing_company": "Лизинговая?",
    "counterparty_main_okved": "ОКВЭД контрагента",
    "llm_level1": "LLM: Направление",
    "llm_level2": "LLM: Категория",
    "llm_level3": "LLM: Подкатегория",
}

# ── Русификация тегов ─────────────────────────────────────
TAG_RU = {
    # Level 1
    "Incoming": "Поступление",
    "Outgoing": "Списание",
    # Level 2 — кредит
    "Customers": "От заказчиков/покупателей",
    "OwnTransfer": "Перевод собственных средств",
    "CashDeposit": "Внесение наличных",
    "Loans": "Займы/кредиты",
    "Refunds": "Возвраты",
    # Level 2 — дебет
    "Suppliers": "Платежи контрагентам",
    "Leasing": "Лизинговые платежи",
    "CreditRepay": "Погашение кредита",
    "LoanRepay": "Погашение займа",
    "CommFinRepay": "Погашение ком. финансирования",
    "LoansIssued": "Займы выданные",
    "CashWithdraw": "Снятие наличных",
    "BankFees": "Банковские комиссии",
    "Taxes": "Отчисления, налоги",
    "Other": "Иное",
    # Level 3
    "Transport": "Транспортные услуги",
    "Platon": "Платон (РТИС)",
    "Fuel": "Топливо, ГСМ, запчасти",
    "Rent": "Аренда",
    "Repair": "Ремонт, ТО",
    "Goods": "Товары",
    "Services": "Услуги",
    "FromFounders": "Займы от учредителей",
    "FromLegal": "Займы от юрлиц",
    "BankCredit": "Банковские кредиты",
    "Factoring": "Факторинг",
    "Accountable": "Подотчётные средства",
    "ErrorPayment": "Ошибочные платежи",
    "CashOut": "Выдача наличных",
    "Affiliated": "Аффилированные переводы",
}


def tr(val: str) -> str:
    """Translate English tag to Russian."""
    return TAG_RU.get(val.strip(), val.strip()) if val else ""


def _apply_header(ws, fields, col_widths):
    """Пишем заголовок со стилями."""
    for col_idx, (field, width) in enumerate(zip(fields, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"


def build_transactions_sheet(ws, rows, fieldnames):
    """Лист «Транзакции» — полный перечень с подсветкой."""
    col_widths = {
        "document_number": 14, "document_operation_date": 14,
        "payer_or_recipient_name": 35, "payer_or_recipient_inn": 15,
        "debit_amount": 18, "credit_amount": 18,
        "payment_purpose": 55, "existing_tag": 25,
        "is_leasing_company": 12, "counterparty_main_okved": 15,
        "llm_level1": 16, "llm_level2": 30, "llm_level3": 25,
    }
    headers = [RU_HEADERS.get(f, f) for f in fieldnames]
    widths = [col_widths.get(f, 15) for f in fieldnames]
    _apply_header(ws, headers, widths)

    for row_idx, row in enumerate(rows, 2):
        level1 = row.get("llm_level1", "")
        row_fill = INCOMING_FILL if level1 == "Incoming" else OUTGOING_FILL if level1 == "Outgoing" else None

        for col_idx, field in enumerate(fieldnames, 1):
            value = row.get(field, "")

            # Русификация
            if field in ("llm_level1", "llm_level2", "llm_level3"):
                value = tr(value)
            if field == "is_leasing_company":
                value = "Да" if value == "True" else "Нет"
            if field in ("debit_amount", "credit_amount"):
                try:
                    value = float(value) if value else 0
                except ValueError:
                    pass

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(field == "payment_purpose"))
            if row_fill:
                cell.fill = row_fill
            if field in ("debit_amount", "credit_amount") and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

    ws.auto_filter.ref = ws.dimensions


def build_summary_sheet(ws, rows):
    """Лист «Сводка» — иерархическая таблица: направление → категория → подкатегория."""
    # Собираем статистику
    stats = defaultdict(lambda: {"count": 0, "debit": 0.0, "credit": 0.0})
    for row in rows:
        l1 = tr(row.get("llm_level1", "")) or "(не классифицировано)"
        l2 = tr(row.get("llm_level2", "")) or "-"
        l3 = tr(row.get("llm_level3", "")) or "-"

        debit = 0.0
        credit = 0.0
        try:
            debit = float(row.get("debit_amount", 0) or 0)
        except ValueError:
            pass
        try:
            credit = float(row.get("credit_amount", 0) or 0)
        except ValueError:
            pass

        key = (l1, l2, l3)
        stats[key]["count"] += 1
        stats[key]["debit"] += debit
        stats[key]["credit"] += credit

    # Заголовки
    headers = ["Направление (level1)", "Категория (level2)", "Подкатегория (level3)",
               "Кол-во транзакций", "Сумма дебет", "Сумма кредит"]
    widths = [22, 35, 30, 20, 22, 22]
    _apply_header(ws, headers, widths)

    # Сортировка: по level1, затем level2, level3
    sorted_keys = sorted(stats.keys(), key=lambda k: (k[0], k[1], k[2]))

    r = 2
    prev_l1 = None
    prev_l2 = None

    # Подытоги по level1
    l1_totals = defaultdict(lambda: {"count": 0, "debit": 0.0, "credit": 0.0})
    for key, s in stats.items():
        l1_totals[key[0]]["count"] += s["count"]
        l1_totals[key[0]]["debit"] += s["debit"]
        l1_totals[key[0]]["credit"] += s["credit"]

    for key in sorted_keys:
        l1, l2, l3 = key
        s = stats[key]

        # Подытог по level1 перед сменой группы
        if prev_l1 is not None and l1 != prev_l1:
            ws.cell(row=r, column=1, value=f"ИТОГО: {prev_l1}").font = TOTAL_FONT
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = TOTAL_FILL
                ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=4, value=l1_totals[prev_l1]["count"]).font = TOTAL_FONT
            ws.cell(row=r, column=5, value=l1_totals[prev_l1]["debit"]).font = TOTAL_FONT
            ws.cell(row=r, column=5).number_format = '#,##0.00'
            ws.cell(row=r, column=6, value=l1_totals[prev_l1]["credit"]).font = TOTAL_FONT
            ws.cell(row=r, column=6).number_format = '#,##0.00'
            r += 1

        show_l1 = l1 if l1 != prev_l1 else ""
        show_l2 = l2 if l2 != prev_l2 or l1 != prev_l1 else ""

        fill = INCOMING_FILL if "Поступление" in l1 else OUTGOING_FILL if "Списание" in l1 else None

        ws.cell(row=r, column=1, value=show_l1)
        ws.cell(row=r, column=2, value=show_l2)
        ws.cell(row=r, column=3, value=l3)
        ws.cell(row=r, column=4, value=s["count"])
        ws.cell(row=r, column=5, value=s["debit"])
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        ws.cell(row=r, column=6, value=s["credit"])
        ws.cell(row=r, column=6).number_format = '#,##0.00'

        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if fill:
                ws.cell(row=r, column=c).fill = fill

        prev_l1 = l1
        prev_l2 = l2
        r += 1

    # Последний подытог
    if prev_l1:
        ws.cell(row=r, column=1, value=f"ИТОГО: {prev_l1}").font = TOTAL_FONT
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = TOTAL_FILL
            ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=4, value=l1_totals[prev_l1]["count"]).font = TOTAL_FONT
        ws.cell(row=r, column=5, value=l1_totals[prev_l1]["debit"]).font = TOTAL_FONT
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        ws.cell(row=r, column=6, value=l1_totals[prev_l1]["credit"]).font = TOTAL_FONT
        ws.cell(row=r, column=6).number_format = '#,##0.00'
        r += 1

    # Общий итог
    r += 1
    total_count = sum(s["count"] for s in stats.values())
    total_debit = sum(s["debit"] for s in stats.values())
    total_credit = sum(s["credit"] for s in stats.values())
    ws.cell(row=r, column=1, value="ОБЩИЙ ИТОГ").font = Font(name="Calibri", size=12, bold=True)
    ws.cell(row=r, column=4, value=total_count).font = Font(name="Calibri", size=12, bold=True)
    ws.cell(row=r, column=5, value=total_debit).font = Font(name="Calibri", size=12, bold=True)
    ws.cell(row=r, column=5).number_format = '#,##0.00'
    ws.cell(row=r, column=6, value=total_credit).font = Font(name="Calibri", size=12, bold=True)
    ws.cell(row=r, column=6).number_format = '#,##0.00'
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        ws.cell(row=r, column=c).border = THIN_BORDER


def build_comparison_sheet(ws, rows):
    """Лист «SQL vs LLM» — сравнение текущего SQL-тега с LLM-категорией."""
    # cross-tab: SQL tag → LLM level2
    cross = defaultdict(lambda: defaultdict(int))
    for row in rows:
        sql_tag = row.get("existing_tag", "") or "(пусто)"
        llm_l2 = tr(row.get("llm_level2", "")) or "(нет)"
        cross[sql_tag][llm_l2] += 1

    sql_tags = sorted(cross.keys())
    llm_cats = sorted({cat for cats in cross.values() for cat in cats})

    # Заголовки
    headers = ["Текущий SQL-тег"] + llm_cats + ["ИТОГО"]
    widths = [30] + [18] * len(llm_cats) + [14]
    _apply_header(ws, headers, widths)

    for r_idx, sql_tag in enumerate(sql_tags, 2):
        ws.cell(row=r_idx, column=1, value=sql_tag)
        ws.cell(row=r_idx, column=1).border = THIN_BORDER
        row_total = 0
        for c_idx, llm_cat in enumerate(llm_cats, 2):
            val = cross[sql_tag].get(llm_cat, 0)
            row_total += val
            cell = ws.cell(row=r_idx, column=c_idx, value=val if val > 0 else "")
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if val > 0:
                cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        ws.cell(row=r_idx, column=len(llm_cats) + 2, value=row_total).border = THIN_BORDER

    ws.auto_filter.ref = ws.dimensions


def main():
    if not os.path.exists(CSV_PATH):
        print(f"Файл не найден: {CSV_PATH}")
        sys.exit(1)

    with open(CSV_PATH, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        rows = list(reader)
        fieldnames = reader.fieldnames

    print(f"Прочитано {len(rows)} строк из {CSV_PATH}")

    # Фильтруем строки без классификации (пустые llm_level1)
    classified = [r for r in rows if r.get("llm_level1", "").strip()]
    skipped = len(rows) - len(classified)
    if skipped:
        print(f"  -> {len(classified)} классифицированы, {skipped} без LLM-тега (будут на листе, но серые)")

    wb = Workbook()

    # Лист 1: Транзакции
    ws1 = wb.active
    ws1.title = "Транзакции"
    build_transactions_sheet(ws1, rows, fieldnames)

    # Лист 2: Сводка
    ws2 = wb.create_sheet("Сводка по категориям")
    build_summary_sheet(ws2, rows)

    # Лист 3: SQL vs LLM
    ws3 = wb.create_sheet("SQL vs LLM")
    build_comparison_sheet(ws3, rows)

    wb.save(XLSX_PATH)
    print(f"Excel сохранён: {XLSX_PATH}")
    print(f"  Лист 1: Транзакции ({len(rows)} строк)")
    print(f"  Лист 2: Сводка по категориям")
    print(f"  Лист 3: SQL vs LLM (сравнение)")


if __name__ == "__main__":
    main()
