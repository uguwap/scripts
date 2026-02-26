import os
import socket
import select
import threading
import time
from datetime import datetime

import paramiko
import pandas as pd
from clickhouse_driver import Client
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SSH_HOST = "doc.ai-referent.ru"
SSH_PORT = 22
SSH_USER = os.getenv("SSH_USER", "tunnel")
SSH_PASSWORD = "R3CiCUFxLhR5bQrGvV4E"

CH_REMOTE_HOST = "10.10.0.4"
CH_REMOTE_PORT = 9000
CH_USER = "i_litvinov"
CH_PASSWORD = "GHO42hfVoC2pdi91ldQp"
CH_DATABASE = "analytic"

DAYS_BACK = 30

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

ssh_client: paramiko.SSHClient | None = None
forward_server: socket.socket | None = None


def forward_tunnel(local_port: int, remote_host: str, remote_port: int, transport: paramiko.Transport):
    srv = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    srv.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    srv.bind(("127.0.0.1", local_port))
    srv.listen(5)
    srv.settimeout(1)
    global forward_server
    forward_server = srv
    while getattr(threading.current_thread(), "running", True):
        try:
            client_sock, _ = srv.accept()
        except socket.timeout:
            continue
        except OSError:
            break
        try:
            chan = transport.open_channel("direct-tcpip", (remote_host, remote_port), client_sock.getpeername())
        except Exception:
            client_sock.close()
            continue
        if chan is None:
            client_sock.close()
            continue
        threading.Thread(target=relay, args=(client_sock, chan), daemon=True).start()


def relay(sock: socket.socket, chan: paramiko.Channel):
    while True:
        r, _, _ = select.select([sock, chan], [], [], 5)
        if sock in r:
            data = sock.recv(65536)
            if not data:
                break
            chan.sendall(data)
        if chan in r:
            data = chan.recv(65536)
            if not data:
                break
            sock.sendall(data)
    chan.close()
    sock.close()


def connect() -> Client:
    global ssh_client
    print(f"SSH -> {SSH_USER}@{SSH_HOST} ...")
    ssh_client = paramiko.SSHClient()
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh_client.connect(SSH_HOST, port=SSH_PORT, username=SSH_USER, password=SSH_PASSWORD, timeout=15)
    transport = ssh_client.get_transport()
    print("SSH OK")

    local_port = 19000
    fwd = threading.Thread(target=forward_tunnel, args=(local_port, CH_REMOTE_HOST, CH_REMOTE_PORT, transport), daemon=True)
    fwd.running = True
    fwd.start()
    time.sleep(0.5)

    print(f"ClickHouse ({CH_DATABASE}) ...")
    client = Client(host="127.0.0.1", port=local_port, database=CH_DATABASE, user=CH_USER, password=CH_PASSWORD,
                    settings={"strings_encoding": "utf-8"})
    client.execute("SELECT 1")
    print("OK\n")
    return client


def disconnect():
    if forward_server:
        forward_server.close()
    if ssh_client:
        ssh_client.close()
    print("SSH закрыт")


def q(client: Client, sql: str) -> list:
    return client.execute(sql)


# ── 1. Сводка ──────────────────────────────────────────────────────
def collect_summary(client: Client) -> pd.DataFrame:
    print("  [1/9] Сводка ...")
    rows = []

    total_ci = q(client, "SELECT count() FROM counterparties_info")[0][0]
    month_ci = q(client, f"SELECT count() FROM counterparties_info WHERE updated_at >= today() - {DAYS_BACK}")[0][0]
    companies = q(client, f"SELECT count() FROM counterparties_info WHERE updated_at >= today() - {DAYS_BACK} AND status NOT IN ('not_found','Person')")[0][0]
    persons_ci = q(client, f"SELECT count() FROM counterparties_info WHERE updated_at >= today() - {DAYS_BACK} AND status = 'Person'")[0][0]
    not_found = q(client, f"SELECT count() FROM counterparties_info WHERE updated_at >= today() - {DAYS_BACK} AND status = 'not_found'")[0][0]

    total_pi = q(client, "SELECT count() FROM person_info")[0][0]
    month_pi = q(client, f"SELECT count() FROM person_info WHERE updated_at >= today() - {DAYS_BACK}")[0][0]

    total_rel = q(client, "SELECT count() FROM relations")[0][0]
    total_edges = q(client, "SELECT count() FROM relations_edges")[0][0]
    month_edges = q(client, f"SELECT count() FROM relations_edges WHERE updated_at >= today() - {DAYS_BACK}")[0][0]

    rows = [
        ("counterparties_info (всего)", total_ci),
        (f"  за {DAYS_BACK} дней — всего", month_ci),
        (f"  за {DAYS_BACK} дней — компании/ИП", companies),
        (f"  за {DAYS_BACK} дней — определены как физлица", persons_ci),
        (f"  за {DAYS_BACK} дней — не найдено (not_found)", not_found),
        ("", ""),
        ("person_info (всего)", total_pi),
        (f"  за {DAYS_BACK} дней", month_pi),
        ("", ""),
        ("relations (всего записей связей)", total_rel),
        ("relations_edges (консолид. рёбра)", total_edges),
        (f"  за {DAYS_BACK} дней (новые рёбра)", month_edges),
    ]
    return pd.DataFrame(rows, columns=["Показатель", "Значение"])


# ── 2. По дням ─────────────────────────────────────────────────────
def collect_daily(client: Client) -> pd.DataFrame:
    print("  [2/9] По дням ...")

    ci_sql = f"""
        SELECT
            toDate(updated_at) as day,
            countIf(status NOT IN ('not_found','Person')) as companies,
            countIf(status = 'Person') as persons_detected,
            countIf(status = 'not_found') as not_found,
            count() as total
        FROM counterparties_info
        WHERE updated_at >= today() - {DAYS_BACK}
        GROUP BY day
        ORDER BY day
    """
    ci_df = pd.DataFrame(q(client, ci_sql),
                         columns=["day", "companies", "persons_detected", "not_found", "total_ci"])

    pi_sql = f"""
        SELECT toDate(updated_at) as day, count() as persons
        FROM person_info
        WHERE updated_at >= today() - {DAYS_BACK}
        GROUP BY day ORDER BY day
    """
    pi_df = pd.DataFrame(q(client, pi_sql), columns=["day", "persons"])

    edges_sql = f"""
        SELECT toDate(updated_at) as day, count() as new_edges
        FROM relations_edges
        WHERE updated_at >= today() - {DAYS_BACK}
        GROUP BY day ORDER BY day
    """
    edges_df = pd.DataFrame(q(client, edges_sql), columns=["day", "new_edges"])

    df = ci_df.merge(pi_df, on="day", how="outer").merge(edges_df, on="day", how="outer")
    df = df.fillna(0)
    int_cols = [c for c in df.columns if c != "day"]
    df[int_cols] = df[int_cols].astype(int)
    df = df.sort_values("day", ascending=False).reset_index(drop=True)

    df.columns = ["Дата", "Компании/ИП", "Физлица (в CI)", "Не найдено", "Всего CI", "Физлица (PI)", "Цепочки (рёбра)"]
    return df


# ── 3. По проектам ─────────────────────────────────────────────────
def collect_projects(client: Client) -> pd.DataFrame:
    print("  [3/9] По проектам ...")
    sql = """
        SELECT
            arm.project_id,
            count(DISTINCT arm.payer_or_recipient_inn) as unique_inns,
            countDistinct(if(ci.status NOT IN ('not_found','Person','') AND ci.id != '', arm.payer_or_recipient_inn, NULL)) as companies,
            countDistinct(if(ci.status = 'Person', arm.payer_or_recipient_inn, NULL)) as persons_detected,
            countDistinct(if(ci.status = 'not_found', arm.payer_or_recipient_inn, NULL)) as not_found,
            countDistinct(if(ci.id != '' AND ci.id IS NOT NULL, arm.payer_or_recipient_inn, NULL)) as loaded_total
        FROM ai_referent_merge arm
        LEFT JOIN counterparties_info ci ON ci.inn = arm.payer_or_recipient_inn
        WHERE arm.payer_or_recipient_inn != ''
        GROUP BY arm.project_id
        ORDER BY unique_inns DESC
    """
    df = pd.DataFrame(q(client, sql),
                      columns=["project_id", "unique_inns", "companies", "persons_detected", "not_found", "loaded_total"])
    df["coverage_%"] = (df["loaded_total"] / df["unique_inns"].replace(0, 1) * 100).round(1)
    return df


# ── 4. Цепочки по проектам ─────────────────────────────────────────
def collect_chains_by_project(client: Client) -> pd.DataFrame:
    print("  [4/9] Цепочки по проектам ...")
    sql = """
        SELECT
            arm.project_id,
            count(DISTINCT arm.payer_or_recipient_inn) as counterparties_with_id,
            count(DISTINCT (arm.debtor_inn, arm.payer_or_recipient_inn)) as possible_pairs
        FROM ai_referent_merge arm
        INNER JOIN counterparties_info ci ON ci.inn = arm.payer_or_recipient_inn AND ci.id != ''
        INNER JOIN counterparties_info ci2 ON ci2.inn = arm.debtor_inn AND ci2.id != ''
        WHERE arm.payer_or_recipient_inn != ''
          AND arm.debtor_inn != ''
          AND arm.debtor_inn != arm.payer_or_recipient_inn
        GROUP BY arm.project_id
        ORDER BY possible_pairs DESC
    """
    return pd.DataFrame(q(client, sql), columns=["project_id", "counterparties_with_id", "possible_pairs"])


# ── 5. Статусы ─────────────────────────────────────────────────────
def collect_statuses(client: Client) -> pd.DataFrame:
    print("  [5/9] Статусы ...")
    sql = """
        SELECT status, count() as cnt,
               min(updated_at) as first_loaded, max(updated_at) as last_loaded
        FROM counterparties_info
        GROUP BY status ORDER BY cnt DESC
    """
    return pd.DataFrame(q(client, sql), columns=["Статус", "Кол-во", "Первая загрузка", "Последняя загрузка"])


# ── 6. Типы связей ─────────────────────────────────────────────────
def collect_relation_types(client: Client) -> pd.DataFrame:
    print("  [6/9] Типы связей ...")
    sql = """
        SELECT relation_type, count() as cnt,
               count(DISTINCT concat(debtor_id, '-', partner_id)) as unique_pairs
        FROM relations GROUP BY relation_type ORDER BY cnt DESC
    """
    return pd.DataFrame(q(client, sql), columns=["Тип связи", "Записей", "Уникальных пар"])


# ── 7. Аудит качества запросов ─────────────────────────────────────
def collect_audit_summary(client: Client) -> pd.DataFrame:
    print("  [7/9] Аудит: сводка качества ...")
    rows = []

    src = q(client, """
        SELECT
            count() as total_rows,
            count(DISTINCT payer_or_recipient_inn) as unique_inns,
            countIf(payer_or_recipient_inn = '') as empty_inn,
            countIf(payer_or_recipient_inn != '' AND length(payer_or_recipient_inn) NOT IN (10, 12)) as bad_length,
            countIf(payer_or_recipient_inn IN ('0000000000','000000000000')) as all_zeros,
            countIf(payer_or_recipient_inn != '' AND match(payer_or_recipient_inn, '^[0-9]+$') = 0) as non_numeric
        FROM ai_referent_merge
    """)[0]
    rows.append(("ai_referent_merge", "", "", "", "", "", ""))
    rows.append(("  Всего транзакций", src[0], "", "", "", "", ""))
    rows.append(("  Уникальных ИНН контрагентов", src[1], "", "", "", "", ""))
    rows.append(("  Пустой ИНН (транзакций)", src[2], f"{src[2]/max(src[0],1)*100:.1f}%", "", "", "", ""))
    rows.append(("  Невалидная длина (не 10/12)", src[3], "", "", "", "", ""))
    rows.append(("  Все нули (0000000000..)", src[4], "", "", "", "", ""))
    rows.append(("  Нечисловые символы", src[5], "", "", "", "", ""))
    rows.append(("", "", "", "", "", "", ""))

    ci = q(client, """
        SELECT
            count() as total,
            countIf(inn = '') as empty_inn,
            countIf(inn != '' AND length(inn) NOT IN (10, 12)) as bad_length,
            countIf(inn IN ('0000000000','000000000000')) as all_zeros,
            countIf(inn != '' AND match(inn, '^[0-9]+$') = 0) as non_numeric,
            countIf(status = 'not_found') as not_found
        FROM counterparties_info
    """)[0]
    rows.append(("counterparties_info (карточки API)", "", "", "", "", "", ""))
    rows.append(("  Всего записей", ci[0], "", "", "", "", ""))
    rows.append(("  Пустой ИНН", ci[1], "", "", "", "", ""))
    rows.append(("  Невалидная длина (не 10/12)", ci[2], "", "", "", "", ""))
    rows.append(("  Все нули", ci[3], "", "", "", "", ""))
    rows.append(("  Нечисловые символы", ci[4], "", "", "", "", ""))
    rows.append(("  not_found (запрос ушёл, ничего не нашли)", ci[5], f"{ci[5]/max(ci[0],1)*100:.1f}%", "", "", "", ""))
    rows.append(("", "", "", "", "", "", ""))

    dupes = q(client, "SELECT count() FROM (SELECT inn FROM counterparties_info GROUP BY inn HAVING count() > 1)")[0][0]
    rows.append(("  Дубли ИНН в counterparties_info", dupes, "", "", "", "", ""))
    rows.append(("", "", "", "", "", "", ""))

    edges = q(client, """
        SELECT
            count() as total,
            countIf(from_inn = '' OR to_inn = '') as empty_pair,
            countIf(from_inn = to_inn) as self_ref
        FROM relations_edges
    """)[0]
    rows.append(("relations_edges (цепочки L5)", "", "", "", "", "", ""))
    rows.append(("  Всего рёбер", edges[0], "", "", "", "", ""))
    rows.append(("  Пустой from/to ИНН", edges[1], "", "", "", "", ""))
    rows.append(("  Самоссылки (from=to)", edges[2], "", "", "", "", ""))

    return pd.DataFrame(rows, columns=["Показатель", "Значение", "% от общего", "", "", "", ""])


def collect_garbage_inns(client: Client) -> pd.DataFrame:
    print("  [8/9] Аудит: мусорные ИНН ...")
    sql = """
        SELECT
            arm.payer_or_recipient_inn as inn,
            length(arm.payer_or_recipient_inn) as len,
            CASE
                WHEN arm.payer_or_recipient_inn IN ('0000000000','000000000000') THEN 'Все нули'
                WHEN length(arm.payer_or_recipient_inn) NOT IN (10,12) THEN concat('Длина=', toString(length(arm.payer_or_recipient_inn)))
                WHEN match(arm.payer_or_recipient_inn, '^[0-9]+$') = 0 THEN 'Нечисловые символы'
                ELSE 'Другое'
            END as problem,
            count() as transactions,
            count(DISTINCT arm.project_id) as projects,
            any(ci.status) as api_status,
            any(arm.payer_or_recipient_name) as example_name
        FROM ai_referent_merge arm
        LEFT JOIN counterparties_info ci ON ci.inn = arm.payer_or_recipient_inn
        WHERE arm.payer_or_recipient_inn != ''
          AND (
              length(arm.payer_or_recipient_inn) NOT IN (10, 12)
              OR arm.payer_or_recipient_inn IN ('0000000000','000000000000')
              OR match(arm.payer_or_recipient_inn, '^[0-9]+$') = 0
          )
        GROUP BY arm.payer_or_recipient_inn
        ORDER BY transactions DESC
        LIMIT 200
    """
    return pd.DataFrame(q(client, sql),
                        columns=["ИНН", "Длина", "Проблема", "Транзакций", "Проектов", "Статус API", "Пример имени"])


def collect_not_found_top(client: Client) -> pd.DataFrame:
    print("  [9/9] Аудит: топ not_found ...")
    sql = """
        SELECT
            ci.inn,
            length(ci.inn) as len,
            arm_cnt,
            arm_projects,
            any_name,
            ci.updated_at
        FROM counterparties_info ci
        LEFT JOIN (
            SELECT
                payer_or_recipient_inn,
                count() as arm_cnt,
                count(DISTINCT project_id) as arm_projects,
                any(payer_or_recipient_name) as any_name
            FROM ai_referent_merge
            GROUP BY payer_or_recipient_inn
        ) arm ON arm.payer_or_recipient_inn = ci.inn
        WHERE ci.status = 'not_found'
        ORDER BY arm_cnt DESC
        LIMIT 200
    """
    return pd.DataFrame(q(client, sql),
                        columns=["ИНН", "Длина", "Транзакций в merge", "Проектов", "Пример имени", "Дата запроса"])


# ── 10. Анализ запросов к API (качество) ────────────────────────────
def collect_api_requests_quality(client: Client) -> pd.DataFrame:
    print("  [10/11] Анализ качества запросов к API ...")
    rows = []

    # Оценка запросов entities/id (по количеству записей в counterparties_info за период)
    id_requests = q(client, f"""
        SELECT
            count() as total_requests_estimated,
            countIf(status = 'not_found') as not_found_requests,
            countIf(status = 'Person') as person_requests,
            countIf(status NOT IN ('not_found', 'Person', '') AND id != '') as successful_requests,
            countIf(id = '' OR id IS NULL) as requests_without_id
        FROM counterparties_info
        WHERE updated_at >= today() - {DAYS_BACK}
    """)[0]
    
    rows.append(("Оценка запросов к API (за последние 30 дней)", "", "", "", "", ""))
    rows.append(("", "", "", "", "", ""))
    rows.append(("Ручка: /api/v1/entities/id (получение ID по ИНН)", "", "", "", "", ""))
    rows.append(("  Всего запросов (оценка по записям в БД)", id_requests[0], "", "", "", ""))
    rows.append(("  Успешных (найдено ID)", id_requests[3], f"{id_requests[3]/max(id_requests[0],1)*100:.1f}%", "", "", ""))
    rows.append(("  not_found (запрос ушёл, ничего не нашли)", id_requests[1], f"{id_requests[1]/max(id_requests[0],1)*100:.1f}%", "", "", ""))
    rows.append(("  Person (определены как физлица)", id_requests[2], f"{id_requests[2]/max(id_requests[0],1)*100:.1f}%", "", "", ""))
    rows.append(("  Без ID (возможно ошибка)", id_requests[4], f"{id_requests[4]/max(id_requests[0],1)*100:.1f}%", "", "", ""))
    rows.append(("", "", "", "", "", ""))

    # Оценка запросов entities/company и entities/ip (по записям с полными данными)
    company_requests = q(client, f"""
        SELECT
            count() as total_cards,
            countIf(status NOT IN ('not_found', 'Person', '') AND id != '' AND names.short_name[1] != '') as complete_cards,
            countIf(status NOT IN ('not_found', 'Person', '') AND (id = '' OR names.short_name[1] = '')) as incomplete_cards
        FROM counterparties_info
        WHERE updated_at >= today() - {DAYS_BACK}
          AND status NOT IN ('not_found', 'Person', '')
    """)[0]
    
    rows.append(("Ручка: /api/v1/entities/company и /api/v1/entities/ip (полные карточки)", "", "", "", "", ""))
    rows.append(("  Всего карточек загружено", company_requests[0], "", "", "", ""))
    rows.append(("  Полных карточек (с именем и ID)", company_requests[1], f"{company_requests[1]/max(company_requests[0],1)*100:.1f}%", "", "", ""))
    rows.append(("  Неполных карточек (без имени или ID)", company_requests[2], f"{company_requests[2]/max(company_requests[0],1)*100:.1f}%", "", "", ""))
    rows.append(("", "", "", "", "", ""))

    # Оценка запросов RelatedEntities/Chains (по relations_edges)
    chains_requests = q(client, f"""
        SELECT
            count() as total_edges,
            countIf(JSONExtractInt(evidence, 'chain_company_len') >= 3) as deep_chains,
            countIf(JSONExtractInt(evidence, 'chain_company_len') < 3 AND JSONExtractInt(evidence, 'chain_company_len') > 0) as short_chains,
            countIf(JSONExtractInt(evidence, 'chain_company_len') = 0 OR JSONExtractInt(evidence, 'chain_company_len') IS NULL) as no_chains
        FROM relations_edges
        WHERE updated_at >= today() - {DAYS_BACK}
    """)[0]
    
    rows.append(("Ручка: /api/v1/RelatedEntities/Chains (цепочки связей)", "", "", "", "", ""))
    rows.append(("  Всего рёбер создано", chains_requests[0], "", "", "", ""))
    rows.append(("  Глубокие цепочки (L3+)", chains_requests[1], f"{chains_requests[1]/max(chains_requests[0],1)*100:.1f}%", "", "", ""))
    rows.append(("  Короткие цепочки (L1-L2)", chains_requests[2], f"{chains_requests[2]/max(chains_requests[0],1)*100:.1f}%", "", "", ""))
    rows.append(("  Без цепочек (прямые связи)", chains_requests[3], f"{chains_requests[3]/max(chains_requests[0],1)*100:.1f}%", "", "", ""))
    rows.append(("", "", "", "", "", ""))

    return pd.DataFrame(rows, columns=["Показатель", "Значение", "%", "", "", ""])


# ── 11. Анализ дублей и мусорных запросов ─────────────────────────────
def collect_duplicates_and_garbage(client: Client):
    print("  [11/11] Анализ дублей и мусорных запросов ...")
    
    # Дубли запросов (один ИНН запрашивался несколько раз)
    dupes_sql = f"""
        SELECT
            inn,
            count() as request_count,
            min(updated_at) as first_request,
            max(updated_at) as last_request,
            groupArray(DISTINCT status) as statuses,
            any(names.short_name[1]) as example_name
        FROM counterparties_info
        WHERE updated_at >= today() - {DAYS_BACK}
        GROUP BY inn
        HAVING count() > 1
        ORDER BY request_count DESC
        LIMIT 100
    """
    
    dupes_df = pd.DataFrame(q(client, dupes_sql),
                           columns=["ИНН", "Кол-во запросов", "Первый запрос", "Последний запрос", "Статусы", "Пример имени"])
    
    # Мусорные ИНН, на которые были сделаны запросы
    garbage_sql = f"""
        SELECT
            ci.inn,
            length(ci.inn) as len,
            CASE
                WHEN ci.inn IN ('0000000000','000000000000') THEN 'Все нули'
                WHEN length(ci.inn) NOT IN (10,12) THEN concat('Длина=', toString(length(ci.inn)))
                WHEN match(ci.inn, '^[0-9]+$') = 0 THEN 'Нечисловые символы'
                ELSE 'Другое'
            END as problem,
            ci.status,
            ci.updated_at,
            count(DISTINCT arm.project_id) as projects,
            any(arm.payer_or_recipient_name) as example_name
        FROM counterparties_info ci
        LEFT JOIN ai_referent_merge arm ON arm.payer_or_recipient_inn = ci.inn
        WHERE ci.updated_at >= today() - {DAYS_BACK}
          AND (
              length(ci.inn) NOT IN (10, 12)
              OR ci.inn IN ('0000000000','000000000000')
              OR match(ci.inn, '^[0-9]+$') = 0
          )
        GROUP BY ci.inn, length(ci.inn), problem, ci.status, ci.updated_at
        ORDER BY ci.updated_at DESC
        LIMIT 200
    """
    
    garbage_df = pd.DataFrame(q(client, garbage_sql),
                            columns=["ИНН", "Длина", "Проблема", "Статус API", "Дата запроса", "Проектов", "Пример имени"])
    
    return dupes_df, garbage_df
    return dupes_df, garbage_df


# ── Стили ──────────────────────────────────────────────────────────
def style_sheet(ws, df: pd.DataFrame):
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(df.columns) + 1):
        max_len = max(
            len(str(df.columns[col_idx - 1])),
            *(len(str(v)) for v in df.iloc[:, col_idx - 1]) if len(df) > 0 else [0],
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 55)

    ws.auto_filter.ref = ws.dimensions


# ── Main ───────────────────────────────────────────────────────────
def main():
    client = connect()

    print("Сбор данных (только SELECT):")
    try:
        summary = collect_summary(client)
        daily = collect_daily(client)
        projects = collect_projects(client)
        chains = collect_chains_by_project(client)
        statuses = collect_statuses(client)
        rel_types = collect_relation_types(client)
        audit_summary = collect_audit_summary(client)
        garbage_inns = collect_garbage_inns(client)
        not_found_top = collect_not_found_top(client)
        api_quality = collect_api_requests_quality(client)
        dupes_df, garbage_requests_df = collect_duplicates_and_garbage(client)
    finally:
        disconnect()

    print("\nСоединение закрыто. Формируем Excel ...\n")

    today_str = datetime.now().strftime("%Y-%m-%d")
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"egrul_report_{today_str}.xlsx")

    sheets = {
        "Сводка": summary,
        "По дням": daily,
        "По проектам": projects,
        "Цепочки по проектам": chains,
        "Статусы": statuses,
        "Типы связей": rel_types,
        "Аудит качества": audit_summary,
        "Мусорные ИНН": garbage_inns,
        "Топ not_found": not_found_top,
        "Качество запросов API": api_quality,
        "Дубли запросов": dupes_df,
        "Мусорные запросы": garbage_requests_df,
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)
            style_sheet(writer.sheets[name], df)

        ws_daily = writer.sheets["По дням"]
        row = len(daily) + 3
        ws_daily.cell(row=row, column=1, value="ИТОГО:").font = Font(bold=True)
        for ci in range(2, len(daily.columns) + 1):
            col_name = daily.columns[ci - 1]
            if col_name != "Дата":
                val = daily[col_name].sum()
                ws_daily.cell(row=row, column=ci, value=int(val)).font = Font(bold=True)

        ws_summary = writer.sheets["Сводка"]
        ws_summary.cell(row=len(summary) + 3, column=1, value=f"Период: {DAYS_BACK} дней")
        ws_summary.cell(row=len(summary) + 4, column=1, value=f"Сгенерировано: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    print(f"Готово: {output_path}")
    for name, df in sheets.items():
        print(f"  {name}: {len(df)} строк")


if __name__ == "__main__":
    main()
