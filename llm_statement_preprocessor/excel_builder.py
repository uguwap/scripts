from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from logger import log_step


def build_debug_sheet(wb: Workbook, operations: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(title="Debug")
    ws.cell(row=1, column=1, value="operation_index")
    ws.cell(row=1, column=2, value="excel_row")
    ws.cell(row=1, column=3, value="source_line")
    ws.cell(row=1, column=4, value="document_operation_date")
    ws.cell(row=1, column=5, value="document_number")
    ws.cell(row=1, column=6, value="debit_amount")
    ws.cell(row=1, column=7, value="credit_amount")
    ws.cell(row=1, column=8, value="payer_or_recipient_name")

    for col_idx in range(1, 9):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)

    for i, op in enumerate(operations, start=1):
        excel_row = 8 + i
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=excel_row)
        ws.cell(row=i + 1, column=3, value=op.get("source_line"))
        ws.cell(row=i + 1, column=4, value=op.get("document_operation_date"))
        ws.cell(row=i + 1, column=5, value=op.get("document_number"))
        ws.cell(row=i + 1, column=6, value=op.get("debit_amount"))
        ws.cell(row=i + 1, column=7, value=op.get("credit_amount"))
        ws.cell(row=i + 1, column=8, value=op.get("payer_or_recipient_name"))


def build_excel_file(
    header: Dict[str, Any],
    operations: List[Dict[str, Any]],
    output_path: str,
    batch_info: str = ""
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Выписка"
    
    ws['A1'] = "наименование банка кредитных организаций"
    ws['B1'] = header.get("debtor_bank_name", "")
    
    ws['A2'] = "выписка по счету"
    ws['B2'] = header.get("debtor_account_number", "")
    
    ws['A3'] = "код валюты"
    ws['B3'] = header.get("currency_code", "643")
    
    ws['A4'] = "клиент:"
    ws['B4'] = header.get("debtor_name", "")
    
    ws['A5'] = "инн"
    ws['B5'] = header.get("debtor_inn", "")
    
    ws.merge_cells('A7:C7')
    ws['A7'] = "Реквизиты документа, на основании которого была совершенна операция по счету"
    ws['A7'].font = Font(bold=True)
    ws['A7'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.merge_cells('D7:G7')
    ws['D7'] = "Реквизиты плательщика/получателя денежных средств"
    ws['D7'].font = Font(bold=True)
    ws['D7'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.merge_cells('H7:I7')
    ws['H7'] = "Сумма операции по счету"
    ws['H7'].font = Font(bold=True)
    ws['H7'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws['J7'] = ""
    
    ws.merge_cells('K7:M7')
    ws['K7'] = "Реквизиты банка плательщика/получателя денежных средств"
    ws['K7'].font = Font(bold=True)
    ws['K7'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    headers_row_2 = [
        "Дата совершения операции (dd.mm.yyyy) или дата проводки",
        "вид (шифр) или ВО",
        "номер или номер документа",
        "наименование/ФИО",
        "ИНН/КИО",
        "КПП",
        "номер счета",
        "По дебету",
        "По кредиту",
        "Назначение платежа",
        "номер корреспондентского счета",
        "наименование",
        "БИК"
    ]
    
    for col_idx, header_text in enumerate(headers_row_2, start=1):
        cell = ws.cell(row=8, column=col_idx, value=header_text)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row_idx, op in enumerate(operations, start=9):
        ws.cell(row=row_idx, column=1, value=op.get("document_operation_date", ""))
        ws.cell(row=row_idx, column=2, value=op.get("document_type_code", ""))
        ws.cell(row=row_idx, column=3, value=op.get("document_number", ""))
        ws.cell(row=row_idx, column=4, value=op.get("payer_or_recipient_name", ""))
        ws.cell(row=row_idx, column=5, value=op.get("payer_or_recipient_inn", ""))
        ws.cell(row=row_idx, column=6, value=op.get("payer_or_recipient_kpp", ""))
        ws.cell(row=row_idx, column=7, value=op.get("account_number", ""))
        ws.cell(row=row_idx, column=8, value=op.get("debit_amount", 0))
        ws.cell(row=row_idx, column=9, value=op.get("credit_amount", 0))
        ws.cell(row=row_idx, column=10, value=op.get("payment_purpose", ""))
        ws.cell(row=row_idx, column=11, value=op.get("correspondent_account_number", ""))
        ws.cell(row=row_idx, column=12, value=op.get("payer_or_recipient_bank", ""))
        ws.cell(row=row_idx, column=13, value=op.get("bank_bik", ""))
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    build_debug_sheet(wb, operations)
    
    wb.save(output_path)
    batch_suffix = f" [{batch_info}]" if batch_info else ""
    log_step("EXCEL_SAVED", path=output_path, operations=len(operations), batch=batch_suffix.strip())


def build_standard_excel(
    header: Dict[str, Any],
    operations: List[Dict[str, Any]],
    output_path: str
) -> None:
    build_excel_file(header, operations, output_path)

